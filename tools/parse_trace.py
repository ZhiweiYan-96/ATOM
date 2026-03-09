#!/usr/bin/env python3
"""
Parse PyTorch profiler trace JSON to extract kernel information.

Usage:
    python parse_trace.py <trace.json> [--layer N]
"""

import json
import gzip
import sys
import bisect
import argparse
import re
from typing import List, Dict, Any, Tuple, Optional
from openpyxl import Workbook

# Modules to filter out (no corresponding GPU kernel in decode)
FILTER_OUT = ["fill_"]

# Sampling-related modules and low-level ops to filter out in prefill
FILTER_OUT_PREFILL = ["aten::", "aiter::gemm_a16w16", "aiter::mixed_sample"]


# =============================================================================
# Utility Functions
# =============================================================================


def load_trace(filepath: str) -> Dict[str, Any]:
    """Load trace JSON file (supports .gz)."""
    opener = gzip.open if filepath.endswith(".gz") else open
    with opener(filepath, "rt", encoding="utf-8") as f:
        return json.load(f)


def is_within(
    child_ts: float, child_dur: float, parent_ts: float, parent_dur: float
) -> bool:
    """Check if child event is within parent's time range."""
    return child_ts >= parent_ts and (child_ts + child_dur) <= (parent_ts + parent_dur)


def is_kernel_launch(name: str) -> bool:
    """Check if name is a kernel launch (contains 'launch' and 'kernel')."""
    n = name.lower()
    return "launch" in n and "kernel" in n


def should_filter(name: str) -> bool:
    """Check if module should be filtered out."""
    return any(f in name for f in FILTER_OUT)


def should_filter_prefill(name: str) -> bool:
    """Check if module should be filtered out in prefill (sampling ops)."""
    return any(f in name for f in FILTER_OUT_PREFILL)


def write_breakdown_xlsx(
    output_xlsx: str,
    rows: List[List[Any]],
    sheet_name: str,
    avg_rows: Optional[List[List[Any]]] = None,
) -> None:
    """
    Write XLSX breakdown with columns:
    cpu_module, gpu_kernel, duration_us, sum per module,
    avg duration_us, avg sum per module.

    The 1st/4th columns are merged for contiguous identical modules.
    AVG columns are appended to the right in the same table.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(
        [
            "cpu_module",
            "gpu_kernel",
            "duration_us",
            "sum per module",
            "avg duration_us",
            "avg sum per module",
        ]
    )

    def build_groups(block_rows: List[List[Any]]) -> List[Tuple[int, int, str, float]]:
        groups: List[Tuple[int, int, str, float]] = []
        i = 0
        while i < len(block_rows):
            mod = block_rows[i][0]
            j = i + 1
            total = float(block_rows[i][2])
            while j < len(block_rows) and block_rows[j][0] == mod:
                total += float(block_rows[j][2])
                j += 1
            groups.append((i, j - 1, mod, total))
            i = j
        return groups

    main_groups = build_groups(rows) if rows else []
    renamed_group_mods = [g[2] for g in main_groups]
    seen_rmsnorm = 0
    for gi, mod in enumerate(renamed_group_mods):
        if isinstance(mod, str) and "rmsnorm" in mod.lower():
            if seen_rmsnorm == 0:
                renamed_group_mods[gi] = "input_layernorm"
            elif seen_rmsnorm == 1:
                renamed_group_mods[gi] = "post_attn_layernorm"
            seen_rmsnorm += 1

    avg_sum_by_row: Dict[int, float] = {}
    if avg_rows:
        avg_groups = build_groups(avg_rows)
        for start, end, _, total in avg_groups:
            for i in range(start, end + 1):
                avg_sum_by_row[i] = total

    data_start_row = ws.max_row + 1
    for gi, (start, end, _, total) in enumerate(main_groups):
        renamed_mod = renamed_group_mods[gi]
        for idx in range(start, end + 1):
            _, kernel, dur = rows[idx]
            avg_dur = (
                float(avg_rows[idx][2]) if avg_rows and idx < len(avg_rows) else ""
            )
            avg_sum = avg_sum_by_row.get(idx, "")
            ws.append([renamed_mod, kernel, dur, total, avg_dur, avg_sum])

    for start, end, _, _ in main_groups:
        if end > start:
            r1 = data_start_row + start
            r2 = data_start_row + end
            ws.merge_cells(start_row=r1, start_column=1, end_row=r2, end_column=1)
            ws.merge_cells(start_row=r1, start_column=4, end_row=r2, end_column=4)
            if avg_rows:
                ws.merge_cells(start_row=r1, start_column=6, end_row=r2, end_column=6)

    total_duration = sum(float(r[2]) for r in rows) if rows else 0.0
    total_avg_duration = sum(float(r[2]) for r in avg_rows) if avg_rows else ""
    ws.append(["TOTAL", "", total_duration, "", total_avg_duration, ""])

    wb.save(output_xlsx)


def _normalize_module_for_avg(name: str) -> str:
    if not isinstance(name, str):
        return str(name)
    return re.sub(r"model\.layers\.\d+\.", "model.layers.*.", name)


def build_avg_rows_from_layers(
    layer_rows_list: List[List[List[Any]]],
    layer_start_idx: int,
    section_name: str,
) -> Optional[List[List[Any]]]:
    """
    Build AVG rows across layers using layer-3 rows as template.
    Returns None if any layer cannot be aligned by (module, kernel) sequence.
    """
    if not layer_rows_list:
        return []

    base = layer_rows_list[0]
    base_sig = [(_normalize_module_for_avg(r[0]), r[1]) for r in base]

    for rel_idx, rows in enumerate(layer_rows_list[1:], start=1):
        sig = [(_normalize_module_for_avg(r[0]), r[1]) for r in rows]
        if sig != base_sig:
            bad_layer = layer_start_idx + rel_idx
            print(
                f"{section_name} avg skipped: layer {bad_layer} does not match layer {layer_start_idx} layout."
            )
            return None

    n = len(layer_rows_list)
    avg_rows: List[List[Any]] = []
    for i, (mod, kernel) in enumerate(base_sig):
        # Keep original module display style from layer_start_idx rows.
        display_mod = base[i][0]
        avg_dur = (
            sum(float(layer_rows_list[layer_idx][i][2]) for layer_idx in range(n)) / n
        )
        avg_rows.append([display_mod, kernel, avg_dur])
    return avg_rows


# =============================================================================
# Optimized Event Index for fast time-range queries
# =============================================================================


class EventIndex:
    """Pre-indexed events for fast time-range queries."""

    def __init__(self, events: List[Dict]):
        # Filter duration events only
        self.duration_events = [e for e in events if e.get("ph") == "X"]
        self.duration_events.sort(key=lambda x: x["ts"])
        self.ts_list = [e["ts"] for e in self.duration_events]

        # Pre-compute kernel launch flags and prefix sum
        self._is_kernel_launch = [
            is_kernel_launch(e.get("name", "")) for e in self.duration_events
        ]
        self._kernel_prefix_sum = [0]
        for is_kl in self._is_kernel_launch:
            self._kernel_prefix_sum.append(
                self._kernel_prefix_sum[-1] + (1 if is_kl else 0)
            )

    def events_in_range(self, start_ts: float, end_ts: float) -> List[Dict]:
        """Get all duration events within [start_ts, end_ts]."""
        left = bisect.bisect_left(self.ts_list, start_ts)
        right = bisect.bisect_right(self.ts_list, end_ts)
        return [
            e
            for e in self.duration_events[left:right]
            if e["ts"] + e.get("dur", 0) <= end_ts
        ]

    def count_kernel_launches_in_range(self, start_ts: float, end_ts: float) -> int:
        """Count kernel launches within time range (fast using prefix sum)."""
        left = bisect.bisect_left(self.ts_list, start_ts)
        right = bisect.bisect_right(self.ts_list, end_ts)
        count = 0
        for i in range(left, right):
            e = self.duration_events[i]
            if e["ts"] + e.get("dur", 0) <= end_ts and self._is_kernel_launch[i]:
                count += 1
        return count

    def get_direct_children(self, parent: Dict) -> List[Dict]:
        """Get direct children of parent event (optimized)."""
        p_ts = parent["ts"]
        p_end = p_ts + parent.get("dur", 0)

        # Get candidates in parent's time range
        candidates = [e for e in self.events_in_range(p_ts, p_end) if e is not parent]

        if not candidates:
            return []

        # Filter to direct children only (not nested in other candidates)
        # Sort by duration descending - larger events are potential parents
        candidates_sorted = sorted(candidates, key=lambda x: -x.get("dur", 0))

        direct = []
        for i, c in enumerate(candidates_sorted):
            c_ts, c_dur = c["ts"], c.get("dur", 0)
            c_end = c_ts + c_dur
            # Check if c is nested inside any larger candidate
            is_nested = False
            for j in range(i):  # Only check larger (earlier in sorted list)
                o = candidates_sorted[j]
                o_ts = o["ts"]
                o_end = o_ts + o.get("dur", 0)
                if c_ts >= o_ts and c_end <= o_end:
                    is_nested = True
                    break
            if not is_nested:
                direct.append(c)

        return sorted(direct, key=lambda x: x["ts"])

    def count_kernel_launches(self, event: Dict) -> int:
        """Count kernel launches within event's time range."""
        e_ts = event["ts"]
        e_end = e_ts + event.get("dur", 0)
        return self.count_kernel_launches_in_range(e_ts, e_end)

    def has_kernel_launch(self, event: Dict) -> bool:
        """Check if event contains any kernel launch."""
        return self.count_kernel_launches(event) > 0


# =============================================================================
# Legacy functions (for prefill compatibility)
# =============================================================================


def find_events(events: List[Dict], name: str, prefix: bool = False) -> List[Dict]:
    """Find all duration events (ph='X') with given name, sorted by time."""
    if prefix:
        result = [
            e
            for e in events
            if e.get("name", "").startswith(name) and e.get("ph") == "X"
        ]
    else:
        result = [e for e in events if e.get("name") == name and e.get("ph") == "X"]
    return sorted(result, key=lambda x: x["ts"])


def get_gpu_kernels(events: List[Dict], start_ts: float) -> List[Dict]:
    """Get GPU kernels (cat='kernel') starting from given timestamp."""
    result = [e for e in events if e.get("cat") == "kernel" and e["ts"] >= start_ts]
    return sorted(result, key=lambda x: x["ts"])


def get_direct_children(parent: Dict, events: List[Dict]) -> List[Dict]:
    """Get direct children of parent event (excluding nested children)."""
    p_ts, p_dur = parent["ts"], parent.get("dur", 0)

    candidates = [
        e
        for e in events
        if e.get("ph") == "X"
        and e is not parent
        and is_within(e.get("ts", 0), e.get("dur", 0), p_ts, p_dur)
    ]

    direct = []
    for c in candidates:
        c_ts, c_dur = c["ts"], c.get("dur", 0)
        is_direct = not any(
            is_within(c_ts, c_dur, o["ts"], o.get("dur", 0))
            for o in candidates
            if o is not c
        )
        if is_direct:
            direct.append(c)

    return sorted(direct, key=lambda x: x["ts"])


def count_kernel_launches(event: Dict, events: List[Dict]) -> int:
    """Count kernel launches within event's subtree."""
    e_ts, e_dur = event["ts"], event.get("dur", 0)
    return sum(
        1
        for e in events
        if e.get("ph") == "X"
        and is_kernel_launch(e.get("name", ""))
        and is_within(e.get("ts", 0), e.get("dur", 0), e_ts, e_dur)
    )


def has_kernel_launch(event: Dict, events: List[Dict]) -> bool:
    """Check if event's subtree contains any kernel launch."""
    return count_kernel_launches(event, events) > 0


# =============================================================================
# Parse Functions
# =============================================================================


def parse_prefill(events: List[Dict], output_xlsx: str, target_layer: int = 3) -> None:
    """
    Parse prefill phase: find the actual prefill event on CPU trace (user_annotation).

    Warmup rule:
    - If only one prefill exists, it is the actual prefill (no warmup).
    - If >=2 prefills exist:
      - If there is a decode_step_bs* event between prefill[0] and prefill[1], prefill[0]
        is treated as warmup and prefill[1] is the actual prefill.
      - Otherwise, prefill[0] is the actual prefill.
    """
    # CPU side prefill/decode annotations.
    # Accept both legacy "prefill" and traced variants like
    # "prefill_bs_1_ctxlens_tensor([417], ...)".
    prefills = [
        e
        for e in events
        if (e.get("name") == "prefill" or e.get("name", "").startswith("prefill_bs_"))
        and e.get("ph") == "X"
        and e.get("cat") == "user_annotation"
    ]
    prefills = sorted(prefills, key=lambda x: x["ts"])

    if not prefills:
        print("No prefill (user_annotation) events found.")
        write_breakdown_xlsx(output_xlsx, [], sheet_name="prefill")
        return

    actual_prefill_idx = 0
    warmup_detected = False

    # Only evaluate warmup when there are at least two prefills.
    if len(prefills) >= 2:
        first = prefills[0]
        second = prefills[1]
        gap_start = first["ts"] + first.get("dur", 0)
        gap_end = second["ts"]

        # If decode_step_bs appears in [gap_start, gap_end], first prefill is warmup.
        has_decode_between = any(
            e.get("ph") == "X"
            and e.get("cat") == "user_annotation"
            and e.get("name", "").startswith("decode_step_bs")
            and gap_start <= e.get("ts", 0) <= gap_end
            for e in events
        )
        if has_decode_between:
            actual_prefill_idx = 1
            warmup_detected = True

    actual_prefill = prefills[actual_prefill_idx]
    print(f"Found {len(prefills)} prefill events (user_annotation)")
    if warmup_detected:
        print("Warmup detected: decode_step_bs found between prefill[0] and prefill[1]")
    else:
        print("No warmup prefill detected by rule, using prefill[0]")
    print(
        f"Using prefill[{actual_prefill_idx}] "
        f"(ts={actual_prefill.get('ts', 0):.0f}, dur={actual_prefill.get('dur', 0):.0f})"
    )

    # Build prefill hierarchy on the same thread as the selected CPU prefill.
    # Using thread affinity is more robust than category-only filtering.
    prefill_tid = actual_prefill.get("tid")
    prefill_pid = actual_prefill.get("pid")
    prefill_hierarchy_events = [
        e
        for e in events
        if e.get("ph") == "X"
        and e.get("tid") == prefill_tid
        and e.get("pid") == prefill_pid
    ]
    # Build index once for fast subtree queries in prefill parsing.
    prefill_idx = EventIndex(prefill_hierarchy_events)
    level1_children = prefill_idx.get_direct_children(actual_prefill)
    print(
        f"Prefill level 1 (same thread pid={prefill_pid}, tid={prefill_tid}): "
        f"{len(level1_children)} nodes"
    )

    # Keep only level2 children that have kernel launch in their subtree.
    launch_level2_items = []
    for l1 in level1_children:
        l1_name = l1.get("name", "<unknown>")
        level2_children = prefill_idx.get_direct_children(l1)
        level2_with_launch = [
            l2 for l2 in level2_children if prefill_idx.has_kernel_launch(l2)
        ]
        for l2 in level2_with_launch:
            launch_level2_items.append(
                {
                    "level1_name": l1_name,
                    "level2_event": l2,
                }
            )

    print(f"Level2 children with kernel launch: {len(launch_level2_items)}")

    # Layer extraction by rmsnorm positions:
    # each layer has 2 rmsnorm modules, layer N starts at rmsnorm index 2*N (0-based).
    TARGET_LAYER = target_layer
    all_norm_indices = [
        i
        for i, item in enumerate(launch_level2_items)
        if "rmsnorm" in item["level2_event"].get("name", "").lower()
    ]
    # Last rmsnorm is final layernorm, not part of transformer layers.
    norm_indices = all_norm_indices[:-1] if len(all_norm_indices) > 0 else []
    print(
        f"Found {len(all_norm_indices)} rmsnorm modules in level2-with-launch rows "
        f"({len(norm_indices)} used for layer split, excluding final layernorm)"
    )

    mod_start = 0
    mod_end = 0
    norm_start_idx = TARGET_LAYER * 2
    norm_end_idx = (TARGET_LAYER + 1) * 2
    final_norm_idx = (
        all_norm_indices[-1] if len(all_norm_indices) > 0 else len(launch_level2_items)
    )
    if norm_start_idx >= len(norm_indices):
        print(
            f"Not enough rmsnorm modules for layer {TARGET_LAYER}, writing empty XLSX"
        )
    else:
        mod_start = norm_indices[norm_start_idx]
        mod_end = (
            norm_indices[norm_end_idx]
            if norm_end_idx < len(norm_indices)
            else final_norm_idx
        )
        print(
            f"Layer {TARGET_LAYER} range by rmsnorm: "
            f"rows [{mod_start}:{mod_end}) from rmsnorm #{norm_start_idx+1} to #{norm_end_idx+1}"
        )
        print(f"Layer {TARGET_LAYER} modules: {mod_end - mod_start}")

    # Build launch->kernel mapping by correlation id.
    # Build launch candidates from current prefill thread/range once.
    runtime_launches = [
        e
        for e in prefill_hierarchy_events
        if e.get("cat") == "cuda_runtime" and is_kernel_launch(e.get("name", ""))
    ]
    runtime_launches.sort(key=lambda x: x.get("ts", 0))
    runtime_launch_ts = [e.get("ts", 0) for e in runtime_launches]

    item_corrs: List[List[Any]] = []
    corr_needed = set()
    for item in launch_level2_items:
        l2 = item["level2_event"]
        l2_start = l2.get("ts", 0)
        l2_end = l2_start + l2.get("dur", 0)

        left = bisect.bisect_left(runtime_launch_ts, l2_start)
        right = bisect.bisect_right(runtime_launch_ts, l2_end)
        launches_in_l2 = runtime_launches[left:right]
        curr_corrs = []
        for launch in launches_in_l2:
            corr = (launch.get("args") or {}).get("correlation")
            if corr is not None:
                corr_needed.add(corr)
                curr_corrs.append(corr)
        item_corrs.append(curr_corrs)

    # Build kernel index only for correlations we actually need.
    kernel_by_corr: Dict[Any, List[Dict]] = {}
    if corr_needed:
        for e in events:
            if e.get("ph") != "X" or e.get("cat") != "kernel":
                continue
            corr = (e.get("args") or {}).get("correlation")
            if corr is None or corr not in corr_needed:
                continue
            kernel_by_corr.setdefault(corr, []).append(e)
        for corr in kernel_by_corr:
            kernel_by_corr[corr].sort(key=lambda x: x.get("ts", 0))

    item_kernels: List[List[Dict[str, Any]]] = []
    for corrs in item_corrs:
        kernels = []
        for corr in corrs:
            for k in kernel_by_corr.get(corr, []):
                kernels.append({"name": k.get("name", "N/A"), "dur": k.get("dur", 0)})
        item_kernels.append(kernels)

    def build_rows_from_item_range(start: int, end: int) -> List[List[Any]]:
        rows = []
        for i in range(start, end):
            item = launch_level2_items[i]
            mod_name = item["level2_event"].get("name", "<unknown>")
            if should_filter_prefill(mod_name):
                continue
            kernels = [k for k in item_kernels[i] if k["name"] not in ("", "N/A")]
            if not kernels:
                continue
            if "moe_forward" in mod_name.lower():
                rows.extend(process_moe_module(mod_name, len(kernels), 0, kernels))
            else:
                for k in kernels:
                    rows.append(
                        [clean_module_name(mod_name, k["name"]), k["name"], k["dur"]]
                    )
        return rows

    # Target layer rows.
    csv_rows = (
        build_rows_from_item_range(mod_start, mod_end)
        if norm_start_idx < len(norm_indices)
        else []
    )
    print(f"Layer {TARGET_LAYER} launch->kernel mapping rows: {len(csv_rows)}")

    print(f"Prefill decode-style CSV rows (after filters): {len(csv_rows)}")

    # AVG rows from layer 3 to last layer.
    avg_rows = None
    avg_layer_rows: List[List[List[Any]]] = []
    avg_start_layer = 3
    layer = avg_start_layer
    while 2 * layer < len(norm_indices):
        s = norm_indices[2 * layer]
        e_idx = 2 * (layer + 1)
        e = norm_indices[e_idx] if e_idx < len(norm_indices) else final_norm_idx
        avg_layer_rows.append(build_rows_from_item_range(s, e))
        layer += 1
    if avg_layer_rows:
        avg_rows = build_avg_rows_from_layers(
            avg_layer_rows, avg_start_layer, "Prefill"
        )
        if avg_rows is not None:
            print(f"Prefill avg rows: {len(avg_rows)}")

    # Write XLSX for prefill.
    write_breakdown_xlsx(output_xlsx, csv_rows, sheet_name="prefill", avg_rows=avg_rows)


def clean_module_name(name: str, mapped_kernel_name: str = "") -> str:
    """Clean and simplify module name."""
    # Runtime launch wrappers should display the actual launched operator name.
    if "hipmodulelaunchkernel" in name.lower() and mapped_kernel_name not in (
        "",
        "N/A",
    ):
        name = mapped_kernel_name

    # Remove 'aiter::' prefix if present
    if name.startswith("aiter::"):
        name = name[7:]  # len('aiter::') == 7

    # Rename based on keywords (rope takes priority)
    name_lower = name.lower()
    if "rope" in name_lower and "cache" in name_lower:
        return "rope & kv_cache"
    if "rope" in name_lower:
        return "rope"
    if "cache" in name_lower and "gemm" not in name_lower:
        return "kv_cache"

    return name


def process_moe_module(
    mod_name: str, kernel_count: int, start_gpu_idx: int, gpu_kernels: List[Dict]
) -> List[List]:
    """
    Process moe_forward module: categorize kernels by name.

    - 'moesort' in kernel name -> moe_sort
    - 'topk' in kernel name -> moe_topk
    - others -> keep original mod_name

    Returns list of [display_name, gpu_kernel_name, gpu_dur] rows.
    """
    rows = []
    for i in range(kernel_count):
        gpu_idx = start_gpu_idx + i
        gpu_kernel_name = "N/A"
        gpu_dur = 0
        if gpu_idx < len(gpu_kernels):
            gpu = gpu_kernels[gpu_idx]
            gpu_kernel_name = gpu.get("name", "N/A")
            gpu_dur = gpu.get("dur", 0)

        # Determine category based on kernel name
        kernel_lower = gpu_kernel_name.lower()
        if "moesort" in kernel_lower:
            category = "moe_sort"
        elif "topk" in kernel_lower:
            category = "moe_topk"
        else:
            category = clean_module_name(mod_name, gpu_kernel_name)

        # Always show category/module name on each row.
        display_name = category
        rows.append([display_name, gpu_kernel_name, gpu_dur])

    return rows


def process_regular_module(
    mod_name: str, kernel_count: int, start_gpu_idx: int, gpu_kernels: List[Dict]
) -> List[List]:
    """Process regular module and show module name on every row."""
    rows = []
    for i in range(kernel_count):
        gpu_idx = start_gpu_idx + i
        gpu_kernel_name = "N/A"
        gpu_dur = 0
        if gpu_idx < len(gpu_kernels):
            gpu = gpu_kernels[gpu_idx]
            gpu_kernel_name = gpu.get("name", "N/A")
            gpu_dur = gpu.get("dur", 0)
        display_name = clean_module_name(mod_name, gpu_kernel_name)
        rows.append([display_name, gpu_kernel_name, gpu_dur])
    return rows


def parse_decode(events: List[Dict], output_xlsx: str, target_layer: int = 3) -> None:
    """
    Parse decode phase: map capture_graph modules to GPU kernels.

    Output CSV columns: cpu_module, gpu_kernel, duration_us
    """
    print("Building event index...")

    # Find GPU-annotated decode_step events (cat='gpu_user_annotation')
    decode_steps = [
        e
        for e in events
        if e.get("name", "").startswith("decode_step")
        and e.get("ph") == "X"
        and e.get("cat") == "gpu_user_annotation"
    ]
    decode_steps = sorted(decode_steps, key=lambda x: x["ts"])

    if not decode_steps:
        print("No decode_step (gpu_user_annotation) events found.")
        return

    # Skip warmup: find first gap > 100ms (warmup/run boundary)
    # Normal decode gaps are < 5ms, so 100ms is safe threshold
    WARMUP_GAP_THRESHOLD = 100000  # 100ms in microseconds
    actual_run_idx = 0
    found_warmup_boundary = False
    for i in range(1, len(decode_steps)):
        gap = decode_steps[i]["ts"] - (
            decode_steps[i - 1]["ts"] + decode_steps[i - 1].get("dur", 0)
        )
        if gap > WARMUP_GAP_THRESHOLD:
            actual_run_idx = i
            found_warmup_boundary = True
            print(f"Warmup/run boundary at [{i-1}]->[{i}], gap={gap/1000:.1f}ms")
            break

    if not found_warmup_boundary:
        print("No warmup detected (no gap > 100ms), using first decode_step")

    first_ds = decode_steps[actual_run_idx]
    first_ds_name = first_ds.get("name", "")
    target_bs: Optional[int] = None
    if "_bs_" in first_ds_name:
        bs = first_ds_name.split("_bs_")[-1]
        target_cg_name = f"capture_graph_bs_{bs}"
        try:
            target_bs = int(bs)
        except ValueError:
            target_bs = None
    else:
        target_cg_name = "capture_graph"

    print(f"First decode_step: {first_ds_name}")
    print(f"Looking for: {target_cg_name}")

    # Find matching capture_graph
    capture_graphs = [
        e for e in events if e.get("name") == target_cg_name and e.get("ph") == "X"
    ]
    if not capture_graphs and target_bs is not None:
        # Prefer the largest capture_graph_bs_K where K < target_bs.
        lower_bs_candidates: List[Tuple[int, Dict[str, Any]]] = []
        for e in events:
            if e.get("ph") != "X":
                continue
            n = e.get("name", "")
            m = re.match(r"^capture_graph_bs_(\d+)$", n)
            if not m:
                continue
            k = int(m.group(1))
            if k < target_bs:
                lower_bs_candidates.append((k, e))
        if lower_bs_candidates:
            best_bs = max(k for k, _ in lower_bs_candidates)
            capture_graphs = sorted(
                [e for k, e in lower_bs_candidates if k == best_bs],
                key=lambda x: x.get("ts", 0),
            )
            print(f"No exact match, using nearest lower capture_graph_bs_{best_bs}")
    if not capture_graphs:
        # Fallback: find any capture_graph
        capture_graphs = [
            e
            for e in events
            if e.get("name", "").startswith("capture_graph") and e.get("ph") == "X"
        ]
        capture_graphs = sorted(capture_graphs, key=lambda x: x["ts"])
        print("No exact match, using first capture_graph")

    if not capture_graphs:
        print("No capture_graph events found.")
        return

    cg = capture_graphs[0]
    print(f"Using: {cg.get('name')}")

    # Build optimized index only for capture_graph's time range
    cg_start = cg["ts"]
    cg_end = cg_start + cg.get("dur", 0)
    cg_events = [
        e
        for e in events
        if e.get("ph") == "X"
        and e.get("ts", 0) >= cg_start
        and e.get("ts", 0) + e.get("dur", 0) <= cg_end
    ]
    print(f"Events in capture_graph: {len(cg_events)}")
    idx = EventIndex(cg_events)

    # Get GPU kernels from first decode_step (within its duration)
    ds1_start = first_ds["ts"]
    ds1_end = ds1_start + first_ds.get("dur", 0)

    gpu_kernels = [
        e
        for e in events
        if e.get("cat") == "kernel" and ds1_start <= e["ts"] <= ds1_end
    ]
    gpu_kernels = sorted(gpu_kernels, key=lambda x: x["ts"])
    print(f"First decode_step (tid={first_ds.get('tid')}): {first_ds_name}")
    print(
        f"  Range: {ds1_start:.0f} ~ {ds1_end:.0f} (dur={first_ds.get('dur', 0):.0f})"
    )
    print(f"  GPU kernels: {len(gpu_kernels)}")

    # Use optimized index for children lookup
    direct_children = idx.get_direct_children(cg)
    kernel_children = [c for c in direct_children if idx.has_kernel_launch(c)]
    print(f"Direct children with kernels: {len(kernel_children)}")

    # Collect all modules with their kernel info
    all_modules = []  # list of (mod_name, kernel_count, start_gpu_idx)
    gpu_idx = 0

    for child in kernel_children:
        child_name = child.get("name", "")
        if should_filter(child_name):
            continue

        # Get sub-children (actual module names)
        sub_children = idx.get_direct_children(child)
        sub_kernel_children = [sc for sc in sub_children if idx.has_kernel_launch(sc)]

        # Determine modules to process
        modules = sub_kernel_children if sub_kernel_children else [child]

        for mod in modules:
            mod_name = mod.get("name", "<unknown>")
            kernel_count = idx.count_kernel_launches(mod)
            all_modules.append((mod_name, kernel_count, gpu_idx))
            gpu_idx += kernel_count

    # Find norm positions (rmsnorm in name)
    all_norm_indices = [
        i for i, (name, _, _) in enumerate(all_modules) if "rmsnorm" in name.lower()
    ]
    # Last rmsnorm is final layernorm, not part of transformer layers.
    norm_indices = all_norm_indices[:-1] if len(all_norm_indices) > 0 else []
    print(
        f"Found {len(all_norm_indices)} norm modules "
        f"({len(norm_indices)} used for layer split, excluding final layernorm)"
    )

    # Extract layer 3 (4th layer, 0-indexed)
    # Each layer has 2 norms, so layer N starts at norm index 2*N
    TARGET_LAYER = target_layer
    norm_start_idx = TARGET_LAYER * 2  # 6 (7th norm, 0-indexed)
    norm_end_idx = (TARGET_LAYER + 1) * 2  # 8 (9th norm, 0-indexed)

    final_norm_idx = (
        all_norm_indices[-1] if len(all_norm_indices) > 0 else len(all_modules)
    )
    if norm_start_idx >= len(norm_indices):
        print(f"Not enough norms for layer {TARGET_LAYER}")
        return

    # Module range for layer 3: from norm_indices[6] to norm_indices[8] (exclusive)
    mod_start = norm_indices[norm_start_idx]
    mod_end = (
        norm_indices[norm_end_idx]
        if norm_end_idx < len(norm_indices)
        else final_norm_idx
    )

    print(
        f"Layer {TARGET_LAYER}: modules [{mod_start}:{mod_end}] (norms at indices {norm_start_idx}, {norm_start_idx+1})"
    )

    def build_rows_for_module_range(start: int, end: int) -> List[List[Any]]:
        rows = []
        for mod_name, kernel_count, start_gpu_idx in all_modules[start:end]:
            if "moe_forward" in mod_name.lower():
                rows.extend(
                    process_moe_module(
                        mod_name, kernel_count, start_gpu_idx, gpu_kernels
                    )
                )
            else:
                rows.extend(
                    process_regular_module(
                        mod_name, kernel_count, start_gpu_idx, gpu_kernels
                    )
                )
        return rows

    # Target layer rows.
    rows = build_rows_for_module_range(mod_start, mod_end)

    # AVG rows from layer 3 to last layer.
    avg_rows = None
    avg_layer_rows: List[List[List[Any]]] = []
    layer = 3
    while 2 * layer < len(norm_indices):
        s = norm_indices[2 * layer]
        e_idx = 2 * (layer + 1)
        e = norm_indices[e_idx] if e_idx < len(norm_indices) else final_norm_idx
        avg_layer_rows.append(build_rows_for_module_range(s, e))
        layer += 1
    if avg_layer_rows:
        avg_rows = build_avg_rows_from_layers(avg_layer_rows, 3, "Decode")
        if avg_rows is not None:
            print(f"Decode avg rows: {len(avg_rows)}")

    # Write XLSX
    write_breakdown_xlsx(output_xlsx, rows, sheet_name="decode", avg_rows=avg_rows)

    print(f"Layer {TARGET_LAYER} modules: {mod_end - mod_start}")
    print(f"XLSX written to: {output_xlsx} ({len(rows)} rows)")


# =============================================================================
# Main
# =============================================================================


def main():
    parser = argparse.ArgumentParser(
        description="Parse PyTorch profiler trace JSON to extract kernel information."
    )
    parser.add_argument("filepath", help="Path to trace JSON or JSON.GZ file")
    parser.add_argument(
        "--layer", type=int, default=3, help="Target layer index (default: 3)"
    )
    args = parser.parse_args()

    if args.layer < 0:
        print("--layer must be >= 0")
        sys.exit(1)

    filepath = args.filepath
    target_layer = args.layer

    print(f"Loading: {filepath}")
    trace = load_trace(filepath)
    events = trace.get("traceEvents", [])
    print(f"Loaded {len(events)} events\n")

    print("=" * 60)
    print("PREFILL ANALYSIS")
    print("=" * 60)
    parse_prefill(events, "prefill_breakdown.xlsx", target_layer=target_layer)

    print("\n" + "=" * 60)
    print("DECODE ANALYSIS")
    print("=" * 60)
    parse_decode(events, "decode_breakdown.xlsx", target_layer=target_layer)


if __name__ == "__main__":
    main()
